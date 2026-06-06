from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List
from uuid import UUID
from datetime import datetime
import csv
import io
from app.database import get_db
from app import models
from app.security import get_current_user
from app.config import MINIO_BUCKET, MINIO_ENDPOINT, s3_client, to_public_foto_url
from app.schemas import item as item_schema
from app.routers.asesmen import run_async_visual_validation, run_async_assessment
import httpx
import logging


logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/v1",
    tags=["3. Data Warga & Bantuan"]
)

@router.post(
    "/import-csv",
    summary="Sinkronisasi data warga dan foto dari file CSV atau Excel (XLSX)"
)
async def import_csv(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    contents = await file.read()
    filename_lower = file.filename.lower()

    reader = []
    if filename_lower.endswith(('.xlsx', '.xls')):
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        sheet = wb.active

        # Ambil header kolom (baris pertama)
        headers = [cell.value for cell in sheet[1]]
        # Ambil data dari baris kedua hingga akhir
        for r in range(2, sheet.max_row + 1):
            row_dict = {}
            row_has_data = False
            for col_idx, header in enumerate(headers):
                if header:
                    val = sheet.cell(row=r, column=col_idx + 1).value
                    if val is not None:
                        # Jika berupa float bernilai bulat (misal KK ending .0), bersihkan ke int
                        if isinstance(val, float) and val.is_integer():
                            val = int(val)
                        row_dict[str(header)] = str(val).strip()
                        row_has_data = True
                    else:
                        row_dict[str(header)] = ""
            if row_has_data:
                reader.append(row_dict)
    else:
        csv_reader = csv.DictReader(io.StringIO(contents.decode("utf-8")))
        reader = list(csv_reader)

    kolom_sah = [c.name for c in models.Keluarga.__table__.columns]
    sukses = 0
    di_skip = 0
    log_foto = []
    keluarga_ids_for_tasks = set()

# Kolom aset yang harus jadi INTEGER (0/1), bukan boolean
    KOLOM_ASET_INT = {
        "kepemilikan_aset",
        "aset_bergerak_tabung_gas", "aset_bergerak_lemari_es", "aset_bergerak_ac",
        "aset_bergerak_pemanas_air", "aset_bergerak_telepon_rumah", "aset_bergerak_tv_datar",
        "aset_bergerak_emas_perhiasan", "aset_bergerak_komputer_laptop_tablet",
        "aset_bergerak_sepeda_motor", "aset_bergerak_sepeda", "aset_bergerak_mobil",
        "aset_bergerak_perahu", "aset_bergerak_kapal_perahu_motor", "aset_bergerak_smartphone",
    }

    # Kolom yang harus jadi INTEGER (Ditambahkan ID material rumah dan PBI)
    KOLOM_INT = {
        "desil_nasional", "jumlah_anggota_keluarga", "jumlah_jenis_usaha",
        "jumlah_pekerja_dibayar", "jumlah_pekerja_tidak_dibayar",
        "luas_lantai_bangunan", "id_dayapenerangan",
        "lahan_tempat_lain", "rumah_tempat_lain",
        "jml_sapi", "jml_kerbau", "jml_kuda", "jml_babi", "jml_kambing_domba",
        "id_status_penguasaan_bangunan", "id_lantai_terluas", "id_dinding_terluas",
        "id_atap_terluas", "id_sumber_airminum", "id_sumberpenerangan", 
        "id_bb_utama", "id_fasilitas_bab", "id_jenis_kloset", "id_pembuangan_tinja",
        "pbi"
    }

    def safe_int(v, default=0):
        try:
            return int(float(v)) if v and str(v).strip() not in ("", "nan") else default
        except (ValueError, TypeError):
            return default

    def fix_nik(v):
        """Konversi scientific notation '3.57301E+15' → '357301XXXXXXX'"""
        if not v:
            return None
        try:
            # Jika berupa scientific notation, konversi ke integer lalu string
            return str(int(float(v)))
        except (ValueError, TypeError):
            return str(v).strip()

    # Mapping dari header CSV/Excel DTKS ke kolom Database
    MAPPING_DTKS = {
        "NIK": "nik",
        "Nik": "nik",
        "no_kk": "no_kk",
        "nama": "nama_kepala_keluarga",
        "pbi": "pbi",
        "id_status_penguasaan_bangunan": "id_status_penguasaan_bangunan",
        "id_lantai_terluas": "id_lantai_terluas",
        "luas_lantai_bangunan": "luas_lantai_bangunan",
        "id_dinding_terluas": "id_dinding_terluas",
        "id_atap_terluas": "id_atap_terluas",
        "id_sumber_airminum": "id_sumber_airminum",
        "id_sumberpenerangan": "id_sumberpenerangan",
        "id_bb_utama": "id_bb_utama",
        "id_fasilitas_bab": "id_fasilitas_bab",
        "id_jenis_kloset": "id_jenis_kloset",
        "id_pembuangan_tinja": "id_pembuangan_tinja",
        "lahan_tempat_lain": "lahan_tempat_lain",
        "rumah_tempat_lain": "rumah_tempat_lain",
        "jml_sapi": "jml_sapi",
        "jml_kerbau": "jml_kerbau",
        "jml_kuda": "jml_kuda",
        "jml_babi": "jml_babi",
        "jml_kambing_domba": "jml_kambing_domba",
        "Foto_Rumah": "url_foto_rumah",
        "foto_rumah": "url_foto_rumah",
        "url_foto_rumah": "url_foto_rumah",
        "Foto_rumah": "url_foto_rumah",
        "FOTO_RUMAH": "url_foto_rumah",
        "foto_rumah_tampak_dalam": "foto_rumah_tampak_dalam",
        "Foto_rumah_tampak_dalam": "foto_rumah_tampak_dalam",
        "Foto_Rumah_Tampak_Dalam": "foto_rumah_tampak_dalam",
        "FOTO_RUMAH_TAMPAK_DALAM": "foto_rumah_tampak_dalam",
    }

    def is_not_null(value) -> bool:
        return value is not None and str(value).strip() not in ("", "nan", "NaN", "None")

    def build_defaults_from_db() -> dict:
        defaults = {}
        for col_name in kolom_sah:
            if col_name in ("id", "no_kk", "nik"):
                continue
            col = getattr(models.Keluarga, col_name)
            if col_name in KOLOM_INT or col_name in KOLOM_ASET_INT:
                avg_val = db.query(func.avg(col)).filter(col.isnot(None)).scalar()
                defaults[col_name] = int(avg_val) if avg_val is not None else 0
            else:
                # Setara dengan: SELECT * FROM keluarga WHERE <col_name> IS NULL
                # (dipakai untuk identifikasi nilai kosong di database)
                mode_row = (
                    db.query(col, func.count(col).label("cnt"))
                    .filter(col.isnot(None))
                    .group_by(col)
                    .order_by(func.count(col).desc())
                    .first()
                )
                defaults[col_name] = mode_row[0] if mode_row else None
        return defaults

    defaults_db = build_defaults_from_db()

    def to_int(value, default=0):
        try:
            if not is_not_null(value):
                return default
            return int(float(value))
        except (TypeError, ValueError):
            return default

    def hitung_skor_bantuan(row: dict) -> dict:
        skor_pkh_plus = 0.0
        skor_aspd = 0.0

        desil = to_int(row.get("desil_nasional", 10), 10)
        if desil in (1, 2):
            skor_pkh_plus += 40.0
        elif desil in (3, 4):
            skor_pkh_plus += 20.0

        lantai = to_int(row.get("id_lantai_terluas", 0), 0)
        if lantai in (7, 8):
            skor_pkh_plus += 20.0

        atap = to_int(row.get("id_atap_terluas", 0), 0)
        if atap in (5, 7):
            skor_pkh_plus += 15.0

        motor = to_int(row.get("aset_bergerak_sepeda_motor", 2), 2)
        if motor == 1:
            skor_pkh_plus -= 15.0

        kolom_disabilitas = [
            "id_penglihatan", "id_pendengaran", "id_berjalan_atau_naik_tangga",
            "id_menggunakan_tangan_jari", "id_belajar_kemampuan_intelektual",
            "id_pengendalian_perilaku", "id_berbicara_komunikasi",
            "id_mengurus_diri", "id_mengingat_berkonsentrasi", "id_kesedihan_depresi"
        ]

        kondisi_terberat = 4
        for col in kolom_disabilitas:
            nilai = to_int(row.get(col, 4), 4)
            if 0 < nilai < kondisi_terberat:
                kondisi_terberat = nilai

        if kondisi_terberat == 1:
            skor_aspd += 95.0
        elif kondisi_terberat == 2:
            skor_aspd += 70.0
        elif kondisi_terberat == 3:
            skor_aspd += 30.0

        return {
            "skor_pkh_plus": max(0.0, min(100.0, float(skor_pkh_plus))),
            "skor_aspd": max(0.0, min(100.0, float(skor_aspd)))
        }

    async with httpx.AsyncClient() as client:
        for idx_row, raw_row in enumerate(reader):
            try:
                row = {}
                for k, v in raw_row.items():
                    if k:
                        cleaned_key = str(k).strip().lstrip("\ufeff")
                        db_key = MAPPING_DTKS.get(cleaned_key, cleaned_key)
                        row[db_key] = v

                if idx_row < 3:
                    log_foto.append(f"[DEBUG] Row {idx_row+1} keys: {list(row.keys())[:10]}")

                no_kk_row = (row.get("no_kk") or row.get("nomor_kartu_keluarga") or "").strip()
                if not no_kk_row or no_kk_row.lower() == "nan":
                    di_skip += 1
                    log_foto.append(
                        f"Baris {idx_row + 1} di-skip: 'no_kk' kosong. Header terdeteksi: {list(row.keys())[:10]}"
                    )
                    continue

                # --- MENGGUNAKAN METODE POP ---
                raw_urls = row.pop("url_foto_rumah", "")
                raw_urls_dalam = row.pop("foto_rumah_tampak_dalam", "")

                # 1. Bersihkan Data Keluarga
                data_bersih = {}
                for k, v in row.items():
                    if k not in kolom_sah:
                        continue

                    val_str = str(v).strip().upper() if v and str(v).strip() not in ("", "nan") else ""

                    if k in KOLOM_ASET_INT:
                        data_bersih[k] = 1 if val_str in ("YA", "1", "TRUE") else 0

                    elif k in ("nik", "no_kk", "nomor_kartu_keluarga"):
                        data_bersih[k] = fix_nik(v)

                    elif k.startswith("kode_"):
                        data_bersih[k] = val_str.replace(".", "") if val_str else None

                    elif k in KOLOM_INT:
                        data_bersih[k] = safe_int(v)

                    else:
                        data_bersih[k] = v if v and str(v).strip() not in ("", "nan") else None

                # 1b. Isi data kosong dari default database
                for col_name, default_val in defaults_db.items():
                    if col_name in data_bersih and not is_not_null(data_bersih[col_name]):
                        if default_val is not None:
                            data_bersih[col_name] = default_val

                # 2. Cek Idempotensi & History
                keluarga_lama = db.query(models.Keluarga).filter(
                    models.Keluarga.no_kk == data_bersih.get("no_kk")
                ).first()

                if keluarga_lama:
                    # Cek apakah ada perubahan variabel
                    any_changes = False
                    for k, v in data_bersih.items():
                        old_val = getattr(keluarga_lama, k, None)
                        if old_val != v:
                            any_changes = True
                            break
                    
                    if not any_changes:
                        di_skip += 1
                        log_foto.append(f"KK {no_kk_row}: DITOLAK (Duplikat, tidak ada perubahan variabel)")
                        continue

                    # Ada perubahan variabel: Arsipkan data lama
                    data_histori = {c.name: getattr(keluarga_lama, c.name) for c in models.Keluarga.__table__.columns}
                    data_histori.pop("id", None)
                    data_histori["id_keluarga"] = keluarga_lama.id

                    arsip_baru = models.KeluargaHistory(**data_histori)
                    db.add(arsip_baru)

                    for k, v in data_bersih.items():
                        setattr(keluarga_lama, k, v)
                    keluarga_diproses = keluarga_lama
                    
                    db.flush()
                else:
                    keluarga_baru = models.Keluarga(**data_bersih)
                    db.add(keluarga_baru)
                    keluarga_diproses = keluarga_baru
                    db.flush()

                # 3. Hitung skor SETELAH data bersih dan terisi
                skor = hitung_skor_bantuan(data_bersih)

                # Tandai status awal sebagai "proses" agar data tidak terlihat sebelum selesai
                hitung = db.query(models.Perhitungan).filter(
                    models.Perhitungan.keluarga_id == keluarga_diproses.id
                ).first()
                if not hitung:
                    hitung = models.Perhitungan(
                        keluarga_id=keluarga_diproses.id,
                        user_id=current_user.id
                    )
                    db.add(hitung)

                hitung.skor_aspd = skor.get("skor_aspd", 0.0)
                hitung.skor_pkh_plus = skor.get("skor_pkh_plus", 0.0)

                is_processing = hitung.status_validasi not in ("diterima", "ditolak")
                if is_processing:
                    hitung.status_validasi = "proses"

                # 3. PROSES URL FOTO (Download ke MinIO)
                for tipe, raw_photo_urls in [("tampak_luar", raw_urls), ("tampak_dalam", raw_urls_dalam)]:
                    if not raw_photo_urls:
                        continue

                    is_tampak_dalam = (tipe == "tampak_dalam")

                    # Idempotency: cek apakah foto dengan tipe ini sudah ada untuk keluarga ini
                    foto_tipe_ada = db.query(models.Foto).filter(
                        models.Foto.keluarga_id == keluarga_diproses.id,
                        models.Foto.tampak_dalam == is_tampak_dalam
                    ).first()

                    if foto_tipe_ada:
                        log_foto.append(f"KK {no_kk_row}: SKIP foto {tipe} (sudah ada di DB)")
                        continue

                    # Bersihkan dan pecah URL
                    cleaned_urls = [u.strip(" []\"'") for u in str(raw_photo_urls).split(",") if u.strip(" []\"'")]
                    
                    for index, original_url in enumerate(cleaned_urls):
                        try:
                            foto_res = await client.get(original_url, follow_redirects=True, timeout=10.0)
                            if foto_res.status_code == 200:
                                nama_file_minio = f"{keluarga_diproses.id}_{tipe}_{index}.jpg"
                                s3_client.put_object(
                                    Bucket=MINIO_BUCKET,
                                    Key=nama_file_minio,
                                    Body=foto_res.content,
                                    ContentType="image/jpeg"
                                )
                                url_minio_final = f"http://{MINIO_ENDPOINT}/{MINIO_BUCKET}/{nama_file_minio}"
                                db.add(models.Foto(
                                    keluarga_id=keluarga_diproses.id,
                                    url_foto=url_minio_final,
                                    sumber="dataset_csv",
                                    nama_file_asli=original_url,
                                    tampak_dalam=(tipe == "tampak_dalam")
                                ))
                                log_foto.append(f"KK {no_kk_row}: BERHASIL upload foto {tipe}")
                            else:
                                log_foto.append(f"KK {no_kk_row}: Gagal download foto {tipe} (HTTP {foto_res.status_code})")
                        except Exception as e:
                            log_foto.append(f"KK {no_kk_row}: ERROR foto {tipe} → {str(e)}")

                sukses += 1
                if is_processing:
                    keluarga_ids_for_tasks.add(keluarga_diproses.id)

            except Exception as e:
                print(f"[ERROR] Baris {idx_row + 1} dengan KK {raw_row.get('nomor_kartu_keluarga', '?')} gagal diproses: {str(e)}")
                db.rollback()
                di_skip += 1
                log_foto.append(f"Error fatal baris KK {raw_row.get('nomor_kartu_keluarga', '?')}: {str(e)}")
                continue

    db.commit()

    for keluarga_id in keluarga_ids_for_tasks:
        background_tasks.add_task(run_async_visual_validation, keluarga_id, current_user.id)
        background_tasks.add_task(run_async_assessment, keluarga_id, current_user.id)

    return {
        "status": "Sukses",
        "pesan": f"{sukses} data keluarga berhasil disinkronisasi, {di_skip} baris dilewati.",
        "log_proses_foto": log_foto
    }

# ENDPOINT MANAJEMEN BANTUAN (FRONTEND)
@router.get(
    "/manajemen-bantuan",
    summary="Ambil data gabungan Keluarga dan Perhitungan AI untuk tabel Manajemen Bantuan",
    response_model=List[item_schema.ManajemenBantuanResponse]
)
async def get_manajemen_bantuan(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    results = db.query(models.Keluarga, models.Perhitungan).outerjoin(
        models.Perhitungan, models.Perhitungan.keluarga_id == models.Keluarga.id
    ).all()
    print(f"[DEBUG] Jumlah keluarga yang di-query untuk manajemen bantuan: {len(results)}")
    response_data = []
    for k, p in results:
        tahap_ui = p.status_validasi if p and p.status_validasi else "analisis"
        rekomendasi_list = p.rekomendasi_bantuan if p and p.rekomendasi_bantuan else []
        bantuan_list = rekomendasi_list if tahap_ui in ("validasi", "diterima", "ditolak") else []
        
        row = item_schema.ManajemenBantuanResponse(
            id_keluarga=str(k.id),
            idLabel=f"ANL-{str(k.id)[:5].upper()}",
            tanggal=datetime.now().strftime("%d %b %Y"),
            nama=k.nama_kepala_keluarga or "-",
            nik=k.nik or k.no_kk or "-",
            wilayah=k.kabupaten_kota or "-",
            kecamatan=k.kecamatan or "-",
            desil=k.desil_nasional or 0,
            skorASPD=p.skor_aspd if p and p.skor_aspd is not None else 0.0,
            skorPKHPlus=p.skor_pkh_plus if p and p.skor_pkh_plus else 0.0,
            tahap=tahap_ui,
            bantuan=bantuan_list,
            rekomendasiBantuan=rekomendasi_list,
            skorKesejahteraan=100.0 - (p.skor_aspd if p and p.skor_aspd is not None else 0.0),
            aiReasoning=p.reasoning_tim3 if p and p.reasoning_tim3 else "Data reasoning belum tersedia dari AI.",
            
            # Mapping variabel dinamis dari database keluarga
            kelurahan_desa=k.kelurahan_desa,
            jumlah_anggota_keluarga=k.jumlah_anggota_keluarga,
            luas_lantai_bangunan=k.luas_lantai_bangunan,
            id_lantai_terluas=k.id_lantai_terluas,
            id_dinding_terluas=k.id_dinding_terluas,
            id_atap_terluas=k.id_atap_terluas,
            id_sumber_airminum=k.id_sumber_airminum,
            id_sumberpenerangan=k.id_sumberpenerangan,
            id_bb_utama=k.id_bb_utama,
            id_fasilitas_bab=k.id_fasilitas_bab,
            id_jenis_kloset=k.id_jenis_kloset,
            id_pembuangan_tinja=k.id_pembuangan_tinja,
            id_disabilitas=k.id_disabilitas,
            tingkat_disabilitas=k.tingkat_disabilitas,
            pbi=k.pbi,
            kpm_jawara=k.kpm_jawara,
            putri_jawara=k.putri_jawara,
            aspd=k.aspd,
            eks_ppks_jawara=k.eks_ppks_jawara,
            ppks_jawara=k.ppks_jawara,
            kemiskinan_ekstrem=k.kemiskinan_ekstrem,
            pkh_plus=k.pkh_plus,
            aset_bergerak_tabung_gas=k.aset_bergerak_tabung_gas,
            aset_bergerak_lemari_es=k.aset_bergerak_lemari_es,
            aset_bergerak_ac=k.aset_bergerak_ac,
            aset_bergerak_pemanas_air=k.aset_bergerak_pemanas_air,
            aset_bergerak_telepon_rumah=k.aset_bergerak_telepon_rumah,
            aset_bergerak_tv_datar=k.aset_bergerak_tv_datar,
            aset_bergerak_emas_perhiasan=k.aset_bergerak_emas_perhiasan,
            aset_bergerak_komputer_laptop_tablet=k.aset_bergerak_komputer_laptop_tablet,
            aset_bergerak_sepeda_motor=k.aset_bergerak_sepeda_motor,
            aset_bergerak_sepeda=k.aset_bergerak_sepeda,
            aset_bergerak_mobil=k.aset_bergerak_mobil,
            aset_bergerak_perahu=k.aset_bergerak_perahu,
            aset_bergerak_kapal_perahu_motor=k.aset_bergerak_kapal_perahu_motor,
            aset_bergerak_smartphone=k.aset_bergerak_smartphone
        )
        response_data.append(row)
        
    return response_data

@router.get(
    "/manajemen-bantuan/{id_keluarga}",
    response_model=item_schema.DetailKeluargaResponse,
    summary="Ambil detail lengkap satu keluarga untuk halaman DetailHasil"
)
async def get_detail_manajemen_bantuan(
    id_keluarga: UUID,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    k = db.query(models.Keluarga).filter(models.Keluarga.id == id_keluarga).first()
    if not k:
        raise HTTPException(status_code=404, detail="Data keluarga tidak ditemukan.")
        
    p = db.query(models.Perhitungan).filter(models.Perhitungan.keluarga_id == id_keluarga).first()
    f = db.query(models.Foto).filter(models.Foto.keluarga_id == id_keluarga).order_by(models.Foto.diunggah_pada.desc()).first()
    fotos = db.query(models.Foto).filter(models.Foto.keluarga_id == id_keluarga).order_by(models.Foto.diunggah_pada.asc()).all()

    if p and p.status_validasi == "proses":
        raise HTTPException(status_code=409, detail="Data masih diproses, silakan coba lagi.")
    
    tahap_ui = p.status_validasi if p and p.status_validasi else "analisis"
    rekomendasi_list = p.rekomendasi_bantuan if p and p.rekomendasi_bantuan else []
    bantuan_list = rekomendasi_list if tahap_ui in ("validasi", "diterima", "ditolak") else []
    
    return item_schema.DetailKeluargaResponse(
        id_keluarga=str(k.id),
        idLabel=f"ANL-{str(k.id)[:5].upper()}",
        tanggal=datetime.now().strftime("%d %b %Y"),
        nama=k.nama_kepala_keluarga or "-",
        nik=k.nik or k.no_kk or "-",
        wilayah=k.kabupaten_kota or "-",
        kecamatan=k.kecamatan or "-",
        desil=k.desil_nasional or 0,
        skorASPD=p.skor_aspd if p and p.skor_aspd is not None else 0.0,
        skorPKHPlus=p.skor_pkh_plus if p and p.skor_pkh_plus else 0.0,
        tahap=tahap_ui,
        bantuan=bantuan_list,
        rekomendasiBantuan=rekomendasi_list,
        skorKesejahteraan=100.0 - (p.skor_aspd if p and p.skor_aspd is not None else 0.0),
        atap=k.id_atap_terluas or 0,
        dinding=k.id_dinding_terluas or 0,
        lantai=k.id_lantai_terluas or 0,
        url_foto=to_public_foto_url(f.url_foto) if f else None,
        foto_urls=[to_public_foto_url(foto.url_foto) for foto in fotos if foto.url_foto],
        visual_match=not p.ada_ketidaksesuaian_visual if p and p.ada_ketidaksesuaian_visual is not None else None,
        visual_reasoning=p.reasoning_tim2 if p else None,
        catatan=p.catatan_petugas if p else None,
        catatan_supervisor=p.catatan_supervisor if p else None,
        aiReasoning=p.reasoning_tim3 if p and p.reasoning_tim3 else "Data reasoning belum tersedia dari AI."
    )

@router.put(
    "/manajemen-bantuan/{id_keluarga}/status",
    summary="Update status validasi dan rekomendasi bantuan",
    response_model=item_schema.DetailKeluargaResponse
)
async def update_status_validasi(
    id_keluarga: UUID,
    request: item_schema.UpdateStatusValidasiRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    p = db.query(models.Perhitungan).filter(models.Perhitungan.keluarga_id == id_keluarga).first()
    if not p:
        p = models.Perhitungan(keluarga_id=id_keluarga)
        db.add(p)
    
    if request.status_validasi:
        p.status_validasi = request.status_validasi
        
    if request.bantuan is not None:
        p.rekomendasi_bantuan = request.bantuan
        
    if request.catatan is not None:
        p.catatan_petugas = request.catatan
        
    if request.catatan_supervisor is not None:
        p.catatan_supervisor = request.catatan_supervisor
        
    db.commit()
    return await get_detail_manajemen_bantuan(id_keluarga, current_user, db)

# ENDPOINT READ DATA (GET)
@router.get(
    "/keluarga",
    summary="Ambil daftar semua keluarga (dengan pagination)"
)
async def list_keluarga(
    skip: int = 0,
    limit: int = 20,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    total = db.query(models.Keluarga).count()
    data = db.query(models.Keluarga).offset(skip).limit(limit).all()
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "data": [item_schema.KeluargaResponse.from_orm(k) for k in data]
    }

@router.get(
    "/keluarga/{keluarga_id}",
    summary="Ambil detail satu keluarga berdasarkan ID"
)
async def get_keluarga(
    keluarga_id: UUID,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    keluarga = db.query(models.Keluarga).filter(models.Keluarga.id == keluarga_id).first()
    if not keluarga:
        raise HTTPException(status_code=404, detail="Keluarga tidak ditemukan")

    foto_terbaru = db.query(models.Foto).filter(
        models.Foto.keluarga_id == keluarga_id
    ).order_by(models.Foto.diunggah_pada.desc()).first()
    
    url_public = foto_terbaru.url_foto if foto_terbaru else None

    return item_schema.KeluargaResponse.from_orm(keluarga)

@router.get(
    "/keluarga/{keluarga_id}/histori",
    summary="Lihat riwayat perubahan asesmen satu keluarga"
)
async def get_histori(
    keluarga_id: UUID,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    logs = db.query(models.LogHistori).filter(
        models.LogHistori.keluarga_id == keluarga_id
    ).order_by(models.LogHistori.timestamp.desc()).all()

    return {
        "keluarga_id": str(keluarga_id),
        "jumlah_riwayat": len(logs),
        "riwayat": [
            {
                "timestamp": log.timestamp,
                "desil_lama": log.desil_lama,
                "desil_baru": log.desil_baru,
                "bantuan_lama": log.bantuan_lama,
                "bantuan_baru": log.bantuan_baru,
            }
            for log in logs
        ]
    }