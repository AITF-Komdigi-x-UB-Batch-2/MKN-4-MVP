import os
import httpx
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, BackgroundTasks
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import datetime, timedelta
from uuid import UUID
import boto3
from botocore.exceptions import ClientError
from dotenv import load_dotenv
import models
import schemas
from database import engine, get_db, SessionLocal
import csv
import io
import uvicorn
import json

# BAGIAN 1: SETUP AWAL
load_dotenv()
models.Base.metadata.create_all(bind=engine)

# BAGIAN 2: KONFIGURASI MINIO
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY")
MINIO_BUCKET = "foto-rumah-warga"

s3_client = boto3.client(
    "s3",
    endpoint_url=f"http://{MINIO_ENDPOINT}",
    aws_access_key_id=MINIO_ACCESS_KEY,
    aws_secret_access_key=MINIO_SECRET_KEY,
)

def ensure_bucket_exists():
    try:
        s3_client.head_bucket(Bucket=MINIO_BUCKET)
    except ClientError:
        s3_client.create_bucket(Bucket=MINIO_BUCKET)
        print(f"[MinIO] Bucket '{MINIO_BUCKET}' berhasil dibuat.")

    policy = {
        "Version": "2012-10-17",
        "Statement": [
            {
                "Effect": "Allow",
                "Principal": "*",
                "Action": ["s3:GetObject"],
                "Resource": [f"arn:aws:s3:::{MINIO_BUCKET}/*"]
            }
        ]
    }
    s3_client.put_bucket_policy(
        Bucket=MINIO_BUCKET,
        Policy=json.dumps(policy)
    )
    print(f"[MinIO] Policy PUBLIC Read-Only berhasil diterapkan pada bucket '{MINIO_BUCKET}'.")

try:
    ensure_bucket_exists()
except Exception as e:
    print(f"[MinIO] Peringatan: Tidak bisa terhubung ke MinIO → {e}")

# BAGIAN 3: KONFIGURASI KEAMANAN JWT
SECRET_KEY = os.getenv("JWT_SECRET_KEY")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

# BAGIAN 4: INISIALISASI APLIKASI FASTAPI
app = FastAPI(
    title="API Pemetaan Kemiskinan Jatim",
    version="2.1",
    description="Backend MVP Tim 4 — Mengorkestrasi alur data dari Tim 1, 2, dan 3."
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def seed_admin_user():
    db = SessionLocal()
    try:
        admin = db.query(models.User).filter_by(username="admin_jatim").first()
        if not admin:
            admin = models.User(
                email="admin@dinsos.go.id",
                username="admin_jatim",
                password_hash=get_password_hash("admin123"),
                role="ADMIN",
                is_active=True
            )
            db.add(admin)
            db.commit()
            print("[Seeder] Akun admin_jatim berhasil dibuat.")
        else:
            # Pastikan akun admin lama punya role ADMIN
            if admin.role != "ADMIN":
                admin.role = "ADMIN"
                db.commit()
            print("[Seeder] Akun admin_jatim sudah tersedia.")
    except Exception as e:
        print(f"[Seeder] Error seeding admin: {e}")
    finally:
        db.close()

AI_BASE_URL = os.getenv("AI_BASE_URL", "http://127.0.0.1:8001")

# BAGIAN 5: DEPENDENCY — PENJAGA PINTU AUTENTIKASI
async def get_current_user(
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
) -> models.User:    

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Token tidak memiliki identitas pengguna.")
    except JWTError:
        raise HTTPException(status_code=401, detail="Token tidak valid atau sudah kadaluarsa.")

    user = db.query(models.User).filter(models.User.username == username).first()
    if not user:
        raise HTTPException(status_code=401, detail="Pengguna tidak ditemukan.")
    return user

# BAGIAN 6: ENDPOINT AUTENTIKASI
@app.post("/auth/register", tags=["Auth"], summary="Registrasi user baru (publik, untuk kebutuhan dev/debug saja)")
def register(payload: schemas.UserCreate, db: Session = Depends(get_db)):

    user_exist = db.query(models.User).filter(models.User.username == payload.username).first()
    
    if user_exist:
        return {
            "status": "Info",
            "pesan": f"Username '{payload.username}' sudah terdaftar. Silakan login ke akun Anda.",
            "action": "login"
        }

    email_exist = db.query(models.User).filter(models.User.email == payload.email).first()
    if email_exist:
        return {
            "status": "Info",
            "pesan": f"Email '{payload.email}' sudah terdaftar. Silakan login ke akun Anda.",
            "action": "login"
        }

    new_user = models.User(
        username=payload.username,
        email=payload.email,
        password_hash=get_password_hash(payload.password),
        role=payload.role
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return {
        "status": "Sukses",
        "pesan": "Akun berhasil dibuat. Silakan login.",
        "username": new_user.username,
        "email": new_user.email,
        "action": "login"
    }

# BAGIAN 6b: MANAJEMEN USER (ADMIN ONLY)
@app.get(
    "/api/v1/users",
    tags=["5. Manajemen User"],
    summary="Ambil daftar semua pengguna sistem (hanya admin)",
    response_model=List[schemas.UserResponse]
)
def get_users(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    users = db.query(models.User).order_by(models.User.dibuat_pada.asc()).all()
    return users

@app.post(
    "/api/v1/users",
    tags=["5. Manajemen User"],
    summary="Buat pengguna baru (hanya admin)",
    response_model=schemas.UserResponse,
    status_code=201
)
def create_user(
    payload: schemas.UserCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Hanya admin yang bisa menambah pengguna baru.")

    if db.query(models.User).filter(models.User.username == payload.username).first():
        raise HTTPException(status_code=409, detail=f"Username '{payload.username}' sudah digunakan.")

    if db.query(models.User).filter(models.User.email == payload.email).first():
        raise HTTPException(status_code=409, detail=f"Email '{payload.email}' sudah terdaftar.")

    # Admin hanya boleh membuat akun ANALIS, bukan ADMIN lain
    if payload.role != "ANALIS":
        raise HTTPException(status_code=403, detail="Admin hanya bisa membuat akun dengan role ANALIS.")

    new_user = models.User(
        username=payload.username,
        email=payload.email,
        password_hash=get_password_hash(payload.password),
        role="ANALIS",
        is_active=True
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.delete(
    "/api/v1/users/{user_id}",
    tags=["5. Manajemen User"],
    summary="Hapus pengguna dari sistem (hanya admin)"
)
def delete_user(
    user_id: UUID,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Hanya admin yang bisa menghapus pengguna.")

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan.")

    # Tidak bisa menghapus akun sendiri
    if str(user.id) == str(current_user.id):
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus akun Anda sendiri.")

    # Tidak bisa menghapus akun ADMIN lain
    if user.role == "ADMIN":
        raise HTTPException(status_code=403, detail="Tidak dapat menghapus akun dengan role ADMIN.")

    db.delete(user)
    db.commit()
    return {"status": "Sukses", "pesan": f"Pengguna '{user.username}' berhasil dihapus."}

@app.patch(
    "/api/v1/users/{user_id}",
    tags=["5. Manajemen User"],
    summary="Update status aktif atau role pengguna (hanya admin)",
    response_model=schemas.UserResponse
)
def update_user(
    user_id: UUID,
    is_active: bool = None,
    role: str = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Hanya admin yang bisa mengubah data pengguna.")

    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan.")

    # Tidak boleh mengubah status diri sendiri via endpoint ini
    if str(user.id) == str(current_user.id):
        raise HTTPException(status_code=400, detail="Gunakan endpoint /api/v1/users/me untuk mengubah profil Anda sendiri.")

    if is_active is not None:
        user.is_active = is_active
    if role is not None:
        # Role hanya bisa diubah ke ANALIS, tidak ke ADMIN
        if role == "ADMIN":
            raise HTTPException(status_code=403, detail="Tidak dapat mengubah role pengguna menjadi ADMIN.")
        if role != "ANALIS":
            raise HTTPException(status_code=400, detail="Role hanya bisa diubah ke ANALIS.")
        user.role = role

    db.commit()
    db.refresh(user)
    return user


@app.post("/auth/login", tags=["Auth"], summary="Login dan dapatkan token JWT")
async def login(
    form: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    
    user = db.query(models.User).filter(models.User.username == form.username).first()
    if not user or not verify_password(form.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Username atau password salah.")
    
    token = create_access_token(data={"sub": user.username})
    return {
        "access_token": token,
        "token_type": "bearer",
        "username": user.username,
        "role": user.role
    }

# BAGIAN 6c: PROFIL DIRI SENDIRI
@app.get(
    "/api/v1/users/me",
    tags=["5. Manajemen User"],
    summary="Ambil profil akun yang sedang login",
    response_model=schemas.UserResponse
)
def get_my_profile(
    current_user: models.User = Depends(get_current_user),
):
    return current_user

@app.put(
    "/api/v1/users/me",
    tags=["5. Manajemen User"],
    summary="Edit profil akun sendiri (username, email, password)",
    response_model=schemas.UserResponse
)
def update_my_profile(
    payload: schemas.UpdateProfileRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Cek konflik username
    if payload.username and payload.username != current_user.username:
        if db.query(models.User).filter(models.User.username == payload.username).first():
            raise HTTPException(status_code=409, detail=f"Username '{payload.username}' sudah digunakan.")
        current_user.username = payload.username

    # Cek konflik email
    if payload.email and payload.email != current_user.email:
        if db.query(models.User).filter(models.User.email == payload.email).first():
            raise HTTPException(status_code=409, detail=f"Email '{payload.email}' sudah terdaftar.")
        current_user.email = payload.email

    # Update password jika diberikan
    if payload.new_password:
        if len(payload.new_password) < 6:
            raise HTTPException(status_code=400, detail="Password baru minimal 6 karakter.")
        current_user.password_hash = get_password_hash(payload.new_password)

    db.commit()
    db.refresh(current_user)

    # Buat token baru jika username berubah
    new_token = None
    if payload.username and payload.username != current_user.username:
        new_token = create_access_token(data={"sub": current_user.username})

    return current_user

def execute_asesmen_sosial_logic(keluarga_id: UUID, user_id: UUID, db: Session):
    keluarga = db.query(models.Keluarga).filter(models.Keluarga.id == keluarga_id).first()
    if not keluarga:
        print(f"[Asinkron] Keluarga {keluarga_id} tidak ditemukan.")
        return
    
    try:
        data_untuk_ai = {
            c.name: getattr(keluarga, c.name)
            for c in models.Keluarga.__table__.columns
        }
        data_untuk_ai.pop("id", None)

        with httpx.Client() as client:
            try:
                response = client.post(
                    f"{AI_BASE_URL}/api/ai/jalur-sosial",
                    json=data_untuk_ai,
                    timeout=30.0
                )
                response.raise_for_status()
                hasil_final = response.json()
            except Exception as e:
                print(f"[Asinkron AI Error] Gagal mendapatkan analisis: {e}")
                return

        rekomendasi_baru = hasil_final.get("rekomendasi_bantuan", [])
        analisis_rag = hasil_final.get("justifikasi_dokumen", "")

        hitung = db.query(models.Perhitungan).filter(
            models.Perhitungan.keluarga_id == keluarga.id
        ).first()

        bantuan_lama = None

        if not hitung:
            hitung = models.Perhitungan(
                keluarga_id=keluarga.id,
                user_id=user_id
            )
            db.add(hitung)
        else:
            bantuan_lama = hitung.rekomendasi_bantuan

        hitung.rekomendasi_bantuan = rekomendasi_baru
        hitung.reasoning_tim3 = analisis_rag
        hitung.skor_aspd = hasil_final.get("skor_aspd", 0.0)
        hitung.skor_pkht = hasil_final.get("skor_pkh_plus", hasil_final.get("skor_pkht", 0.0))
        hitung.status_validasi = "analisis"

        log = models.LogHistori(
            keluarga_id=keluarga.id,
            user_id=user_id,
            desil_lama=None,
            desil_baru=None,
            bantuan_lama=bantuan_lama,
            bantuan_baru=rekomendasi_baru
        )
        db.add(log)
        db.commit()
        print(f"[Asinkron] Asesmen sukses untuk KK {keluarga.nomor_kartu_keluarga}")
    except Exception as e:
        db.rollback()
        print(f"[Asinkron DB Error] {e}")

def run_async_assessment(keluarga_id: UUID, user_id: UUID):
    db_gen = get_db()
    db = next(db_gen)
    try:
        execute_asesmen_sosial_logic(keluarga_id, user_id, db)
    finally:
        db.close()

@app.post(
    "/api/v1/import-csv",
    tags=["1. Import Master Data"],
    summary="Sinkronisasi data warga dan foto dari file CSV atau Excel (XLSX)"
)
async def import_csv(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    contents = await file.read()
    filename_lower = file.filename.lower()

    reader = []
    if filename_lower.endswith(('.xlsx', '.xls')):
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        sheet = wb.active

        # Ambil header kolom (baris pertama)
        headers = [cell.value for cell in sheet[1]]
        # Ambil data dari baris kedua hingga akhir
        for r in range(2, sheet.max_row + 1):
            row_dict = {}
            row_has_data = False
            for col_idx, header in enumerate(headers):
                if header:
                    val = sheet.cell(row=r, column=col_idx + 1).value
                    if val is not None:
                        # Jika berupa float bernilai bulat (misal KK ending .0), bersihkan ke int
                        if isinstance(val, float) and val.is_integer():
                            val = int(val)
                        row_dict[str(header)] = str(val).strip()
                        row_has_data = True
                    else:
                        row_dict[str(header)] = ""
            if row_has_data:
                reader.append(row_dict)
    else:
        # Jalankan parser CSV default
        csv_reader = csv.DictReader(io.StringIO(contents.decode("utf-8")))
        reader = list(csv_reader)

    kolom_sah = [c.name for c in models.Keluarga.__table__.columns]
    sukses = 0
    di_skip = 0
    log_foto = []

    # Kolom aset yang harus jadi INTEGER (0/1), bukan boolean
    KOLOM_ASET_INT = {
        "kepemilikan_aset",
        "aset_bergerak_tabung_gas", "aset_bergerak_lemari_es", "aset_bergerak_ac",
        "aset_bergerak_pemanas_air", "aset_bergerak_telepon_rumah", "aset_bergerak_tv_datar",
        "aset_bergerak_emas_perhiasan", "aset_bergerak_komputer_laptop_tablet",
        "aset_bergerak_sepeda_motor", "aset_bergerak_sepeda", "aset_bergerak_mobil",
        "aset_bergerak_perahu", "aset_bergerak_kapal_perahu_motor", "aset_bergerak_smartphone",
    }

    # Kolom yang harus jadi INTEGER
    KOLOM_INT = {
        "desil_nasional", "jumlah_anggota_keluarga", "jumlah_jenis_usaha",
        "jumlah_pekerja_dibayar", "jumlah_pekerja_tidak_dibayar",
        "luas_lantai_bangunan", "id_dayapenerangan",
        "lahan_tempat_lain", "rumah_tempat_lain",
        "jml_sapi", "jml_kerbau", "jml_kuda", "jml_babi", "jml_kambing_domba",
    }

    def safe_int(v, default=0):
        try:
            return int(float(v)) if v and str(v).strip() not in ("", "nan") else default
        except (ValueError, TypeError):
            return default

    def fix_nik(v):
        """Konversi scientific notation '3.57301E+15' → '357301XXXXXXX'"""
        if not v:
            return None
        try:
            # Jika berupa scientific notation, konversi ke integer lalu string
            return str(int(float(v)))
        except (ValueError, TypeError):
            return str(v).strip()

    # Mapping dari header CSV/Excel DTKS Alvin ke kolom Database
    MAPPING_DTKS = {
        "no_kk": "nomor_kartu_keluarga",
        "nama": "nama_anggota_keluarga",
        "pbi": "pbi_nas",
        "id_status_penguasaan_bangunan": "status_kepemilikan_rumah",
        "id_lantai_terluas": "jenis_lantai_terluas",
        "luas_lantai_bangunan": "luas_lantai",
        "id_dinding_terluas": "jenis_dinding_terluas",
        "id_atap_terluas": "jenis_atap_terluas",
        "id_sumber_airminum": "sumber_air_minum_utama",
        "id_sumberpenerangan": "sumber_penerangan_utama",
        "id_bb_utama": "bahan_bakar_utama_memasak",
        "id_fasilitas_bab": "fasilitas_bab",
        "id_jenis_kloset": "jenis_kloset",
        "id_pembuangan_tinja": "pembuangan_akhir_tinja",
        "lahan_tempat_lain": "aset_tidak_bergerak_lahan_lainnya",
        "rumah_tempat_lain": "aset_tidak_bergerak_rumah_lainnya",
        "jml_sapi": "jumlah_ternak_sapi",
        "jml_kerbau": "jumlah_ternak_kerbau",
        "jml_kuda": "jumlah_ternak_kuda",
        "jml_babi": "jumlah_ternak_babi",
        "jml_kambing_domba": "jumlah_ternak_kambing_domba",
        "Foto_Rumah": "url_foto_rumah",
        "foto_rumah": "url_foto_rumah",
        "url_foto_rumah": "url_foto_rumah",
        "Foto_rumah": "url_foto_rumah",
        "FOTO_RUMAH": "url_foto_rumah",
        "lokasi_foto_rumah": "url_foto_rumah",
        "Lokasi_Foto_Rumah": "url_foto_rumah",
        "LOKASI_FOTO_RUMAH": "url_foto_rumah",
        "foto_rumah_tampak_dalam": "foto_rumah_tampak_dalam",
        "Foto_rumah_tampak_dalam": "foto_rumah_tampak_dalam",
        "Foto_Rumah_Tampak_Dalam": "foto_rumah_tampak_dalam",
        "FOTO_RUMAH_TAMPAK_DALAM": "foto_rumah_tampak_dalam",
        "lokasi_foto_rumah_tampak_dalam": "foto_rumah_tampak_dalam",
        "Lokasi_Foto_Rumah_Tampak_Dalam": "foto_rumah_tampak_dalam",
        "LOKASI_FOTO_RUMAH_TAMPAK_DALAM": "foto_rumah_tampak_dalam"
    }

    async with httpx.AsyncClient() as client:
        for idx_row, raw_row in enumerate(reader):
            try:
                # Terjemahkan key dari format CSV Alvin ke format DB menggunakan MAPPING_DTKS
                row = {}
                for k, v in raw_row.items():
                    if k:
                        db_key = MAPPING_DTKS.get(str(k).strip(), str(k).strip())
                        row[db_key] = v

                no_kk_row = row.get("nomor_kartu_keluarga")
                if not no_kk_row:
                    di_skip += 1
                    log_foto.append(f"Baris {idx_row + 1} di-skip: 'nomor_kartu_keluarga' kosong. Header terdeteksi: {list(row.keys())[:5]}")
                    continue

                # --- MENGGUNAKAN METODE POP ---
                # Mengambil URL dari kedua kolom foto dan menghapusnya dari dictionary row
                raw_urls = row.pop("url_foto_rumah", "")
                raw_urls_dalam = row.pop("foto_rumah_tampak_dalam", "")

                # 1. Bersihkan Data Keluarga
                data_bersih = {}
                for k, v in row.items():
                    if k not in kolom_sah:
                        continue

                    val_str = str(v).strip().upper() if v and str(v).strip() not in ("", "nan") else ""

                    # BUG FIX #1: aset → INTEGER (0/1), bukan bool
                    if k in KOLOM_ASET_INT:
                        data_bersih[k] = 1 if val_str in ("YA", "1", "TRUE") else 0

                    # BUG FIX #2: nik & no_kk → string bersih dari scientific notation
                    elif k in ("nik", "no_kk"):
                        data_bersih[k] = fix_nik(v)

                    elif k.startswith("kode_"):
                        data_bersih[k] = val_str.replace(".", "") if val_str else None

                    elif k in KOLOM_INT:
                        data_bersih[k] = safe_int(v)

                    else:
                        data_bersih[k] = v if v and str(v).strip() not in ("", "nan") else None

                # 2. Cek Idempotensi & History
                keluarga_lama = db.query(models.Keluarga).filter(
                    models.Keluarga.no_kk == data_bersih.get("no_kk")
                ).first()

                if keluarga_lama:
                    # Cek apakah ada perubahan variabel
                    any_changes = False
                    for k, v in data_bersih.items():
                        old_val = getattr(keluarga_lama, k, None)
                        if old_val != v:
                            any_changes = True
                            break
                    
                    if not any_changes:
                        di_skip += 1
                        log_foto.append(f"KK {no_kk_row}: DITOLAK (Duplikat, tidak ada perubahan variabel)")
                        continue

                    # Ada perubahan variabel: Arsipkan data lama
                    data_histori = {c.name: getattr(keluarga_lama, c.name) for c in models.Keluarga.__table__.columns}
                    data_histori.pop("id", None)
                    data_histori["keluarga_id"] = keluarga_lama.id

                    arsip_baru = models.KeluargaHistory(**data_histori)
                    db.add(arsip_baru)

                    for k, v in data_bersih.items():
                        setattr(keluarga_lama, k, v)
                    keluarga_diproses = keluarga_lama
                    
                    db.flush() # Amankan ID sebelum memanggil background task

                    # Jadwalkan ulang analisis sosial di latar belakang
                    background_tasks.add_task(
                        run_async_assessment,
                        keluarga_diproses.id,
                        current_user.id
                    )
                else:
                    keluarga_baru = models.Keluarga(**data_bersih)
                    db.add(keluarga_baru)
                    keluarga_diproses = keluarga_baru

                    db.flush() # Amankan ID sebelum memanggil background task

                    # Jadwalkan analisis sosial pertama kali untuk keluarga baru
                    background_tasks.add_task(
                        run_async_assessment,
                        keluarga_diproses.id,
                        current_user.id
                    )

                # 3. PROSES URL FOTO (Download ke MinIO)
                for tipe, raw_photo_urls in [("tampak_luar", raw_urls), ("tampak_dalam", raw_urls_dalam)]:
                    if not raw_photo_urls:
                        continue

                    # Bersihkan dan pecah URL
                    cleaned_urls = [u.strip(" []\"'") for u in str(raw_photo_urls).split(",") if u.strip(" []\"'")]
                    
                    for index, original_url in enumerate(cleaned_urls):
                        try:
                            foto_ada = db.query(models.Foto).filter(
                                models.Foto.keluarga_id == keluarga_diproses.id,
                                models.Foto.nama_file_asli == original_url
                            ).first()

                            if not foto_ada:
                                foto_res = await client.get(original_url, follow_redirects=True, timeout=10.0)
                                if foto_res.status_code == 200:
                                    nama_file_minio = f"{keluarga_diproses.id}_{tipe}_{index}.jpg"
                                    s3_client.put_object(
                                        Bucket=MINIO_BUCKET,
                                        Key=nama_file_minio,
                                        Body=foto_res.content,
                                        ContentType="image/jpeg"
                                    )
                                    url_minio_final = f"http://{MINIO_ENDPOINT}/{MINIO_BUCKET}/{nama_file_minio}"
                                    db.add(models.Foto(
                                        keluarga_id=keluarga_diproses.id,
                                        url_foto=url_minio_final,
                                        sumber="dataset_csv",
                                        nama_file_asli=original_url,
                                        tampak_dalam=(tipe == "tampak_dalam")
                                    ))
                                    log_foto.append(f"KK {no_kk_row}: BERHASIL upload foto {tipe}")
                                else:
                                    log_foto.append(f"KK {no_kk_row}: Gagal download foto {tipe} (HTTP {foto_res.status_code})")
                        except Exception as e:
                            log_foto.append(f"KK {no_kk_row}: ERROR foto {tipe} → {str(e)}")

                sukses += 1

            except Exception as e:
                # BUG FIX: Rollback session agar baris berikutnya tidak kena PendingRollbackError
                db.rollback()
                di_skip += 1
                log_foto.append(f"Error fatal baris KK {raw_row.get('nomor_kartu_keluarga', '?')}: {str(e)}")
                continue

    db.commit()

    return {
        "status": "Sukses",
        "pesan": f"{sukses} data keluarga berhasil disinkronisasi, {di_skip} baris dilewati.",
        "log_proses_foto": log_foto
    }


# BAGIAN 8: ASESMEN SOSIAL — TIM 1 & TIM 3
@app.post(
    "/api/v1/asesmen/sosial",
    tags=["2. Asesmen Tim 1 & 3"],
    summary="Analisis Tim 1 yang diteruskan ke Tim 3"
)
async def asesmen_sosial(
    payload: schemas.TriggerAsesmenRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    keluarga = db.query(models.Keluarga).filter(
        models.Keluarga.id == payload.keluarga_id
    ).first()
    if not keluarga:
        raise HTTPException(status_code=404, detail="Data keluarga tidak ditemukan.")

    execute_asesmen_sosial_logic(payload.keluarga_id, current_user.id, db)

    # Ambil data terupdate untuk direspon
    hitung = db.query(models.Perhitungan).filter(
        models.Perhitungan.keluarga_id == keluarga.id
    ).first()

    rekomendasi_baru = hitung.rekomendasi_bantuan if hitung else []
    analisis_rag = hitung.reasoning_tim3 if hitung else ""

    return {
        "status": "Sukses",
        "nomor_kk": keluarga.nomor_kartu_keluarga,
        "hasil_rekomendasi_final": rekomendasi_baru,
        "justifikasi_dokumen": analisis_rag
    }

# BAGIAN 9: ASESMEN VISUAL — TIM 2
@app.post(
    "/api/v1/asesmen/visual/{id_keluarga}",
    tags=["3. Asesmen Tim 2"],
    summary="Trigger AI Visual mengirim URL dan ID ke Tim 2"
)
async def asesmen_visual(
    id_keluarga: UUID,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # 1. Cek Data Keluarga & Foto
    keluarga = db.query(models.Keluarga).filter(models.Keluarga.id == id_keluarga).first()
    if not keluarga:
        raise HTTPException(status_code=404, detail="Data keluarga tidak ditemukan.")
    
    # 2. Cari Foto Tampak Luar & Tampak Dalam terbaru di Database
    foto_tampak_luar = db.query(models.Foto).filter(
        models.Foto.keluarga_id == id_keluarga,
        models.Foto.tampak_dalam == False
    ).order_by(models.Foto.diunggah_pada.desc()).first()

    foto_tampak_dalam = db.query(models.Foto).filter(
        models.Foto.keluarga_id == id_keluarga,
        models.Foto.tampak_dalam == True
    ).order_by(models.Foto.diunggah_pada.desc()).first()

    # Fallback ke foto apa saja jika tampak luar tidak ada
    foto_utama = foto_tampak_luar if foto_tampak_luar else db.query(models.Foto).filter(
        models.Foto.keluarga_id == id_keluarga
    ).order_by(models.Foto.diunggah_pada.desc()).first()

    if not foto_utama:
        raise HTTPException(status_code=400, detail="Foto rumah tidak ditemukan untuk keluarga ini.")

    url_foto_luar = foto_tampak_luar.url_foto if foto_tampak_luar else foto_utama.url_foto
    url_foto_dalam = foto_tampak_dalam.url_foto if foto_tampak_dalam else None

    # 3. Siapkan Payload JSON untuk Visual Validator
    payload_ke_tim2 = {
        "image_url": url_foto_luar,
        "foto_rumah_tampak_dalam": url_foto_dalam, 
        "konteks_rumah": {
            "jenis_lantai_terluas": keluarga.jenis_lantai_terluas,
            "jenis_dinding_terluas": keluarga.jenis_dinding_terluas,
            "jenis_atap_terluas": keluarga.jenis_atap_terluas,
        }
    }

    try:
        # 4. Tembak Endpoint Tim 2
        async with httpx.AsyncClient() as client:
            res_ai = await client.post(
                f"{AI_BASE_URL}/api/ai/visual-validator",
                json=payload_ke_tim2,
                timeout=30.0
            )
            res_ai.raise_for_status() 
            hasil_validator = res_ai.json() 

        is_match = hasil_validator.get("is_match", True)
        reasoning = hasil_validator.get("reasoning", "Validasi visual selesai.")

        # 5. Tangkap Response & Simpan ke Perhitungan
        hitung = db.query(models.Perhitungan).filter(
            models.Perhitungan.keluarga_id == keluarga.id
        ).first()
        
        if not hitung:
            hitung = models.Perhitungan(keluarga_id=keluarga.id, user_id=current_user.id)
            db.add(hitung)

        hitung.ada_ketidaksesuaian_visual = not is_match
        hitung.reasoning_tim2 = reasoning
        hitung.foto_id_digunakan = foto_utama.id
        hitung.status_validasi = "validasi"
        
        db.commit()

        return {
            "status": "Sukses",
            "validation": {
                "is_match": is_match,
                "reasoning": reasoning
            },
            "url_foto_divalidasi": foto_utama.url_foto,
            "hasil_tim2": hasil_validator
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Kesalahan koneksi ke Tim 2: {str(e)}")

# BAGIAN 10: ENDPOINT MANAJEMEN BANTUAN (FRONTEND)
@app.get(
    "/api/v1/manajemen-bantuan",
    tags=["4. Read Data"],
    summary="Ambil data gabungan Keluarga dan Perhitungan AI untuk tabel Manajemen Bantuan",
    response_model=List[schemas.ManajemenBantuanResponse]
)
async def get_manajemen_bantuan(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    results = db.query(models.Keluarga, models.Perhitungan).outerjoin(
        models.Perhitungan, models.Keluarga.id == models.Perhitungan.keluarga_id
    ).all()

    response_data = []
    for k, p in results:
        tahap_ui = p.status_validasi if p and p.status_validasi else "analisis"
        bantuan_list = p.rekomendasi_bantuan if p and p.rekomendasi_bantuan else []
        
        row = schemas.ManajemenBantuanResponse(
            id_keluarga=str(k.id),
            idLabel=f"ANL-{str(k.id)[:5].upper()}",
            tanggal=datetime.now().strftime("%d %b %Y"),
            nama=k.nama_anggota_keluarga or "-",
            nik=k.nomor_kartu_keluarga or "-",
            wilayah=k.kabupaten_kota or "-",
            kecamatan=k.kecamatan or "-",
            desil=k.desil_nasional or 0,
            skorASPD=p.skor_aspd if p and p.skor_aspd else 0.0,
            skorPKHT=p.skor_pkht if p and p.skor_pkht else 0.0,
            tahap=tahap_ui,
            bantuan=bantuan_list,
            rekomendasiBantuan=bantuan_list,
            skorKesejahteraan=100.0 - (p.skor_aspd if p and p.skor_aspd else 0.0), # Dummy inversion for sorting
            aiReasoning=p.reasoning_tim3 if p and p.reasoning_tim3 else "Data reasoning belum tersedia dari AI."
        )
        response_data.append(row)
        
    return response_data

@app.get(
    "/api/v1/manajemen-bantuan/{id_keluarga}",
    response_model=schemas.DetailKeluargaResponse,
    tags=["4. Read Data"],
    summary="Ambil detail lengkap satu keluarga untuk halaman DetailHasil"
)
async def get_detail_manajemen_bantuan(
    id_keluarga: UUID,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    k = db.query(models.Keluarga).filter(models.Keluarga.id == id_keluarga).first()
    if not k:
        raise HTTPException(status_code=404, detail="Data keluarga tidak ditemukan.")
        
    p = db.query(models.Perhitungan).filter(models.Perhitungan.keluarga_id == id_keluarga).first()
    f = db.query(models.Foto).filter(models.Foto.keluarga_id == id_keluarga).order_by(models.Foto.diunggah_pada.desc()).first()
    fotos = db.query(models.Foto).filter(models.Foto.keluarga_id == id_keluarga).order_by(models.Foto.diunggah_pada.asc()).all()
    
    tahap_ui = p.status_validasi if p and p.status_validasi else "analisis"
    bantuan_list = p.rekomendasi_bantuan if p and p.rekomendasi_bantuan else []
    
    return schemas.DetailKeluargaResponse(
        id_keluarga=str(k.id),
        idLabel=f"ANL-{str(k.id)[:5].upper()}",
        tanggal=datetime.now().strftime("%d %b %Y"),
        nama=k.nama_anggota_keluarga or "-",
        nik=k.nomor_kartu_keluarga or "-",
        wilayah=k.kabupaten_kota or "-",
        kecamatan=k.kecamatan or "-",
        desil=k.desil_nasional or 0,
        skorASPD=p.skor_aspd if p and p.skor_aspd else 0.0,
        skorPKHT=p.skor_pkht if p and p.skor_pkht else 0.0,
        tahap=tahap_ui,
        bantuan=bantuan_list,
        rekomendasiBantuan=bantuan_list,
        skorKesejahteraan=100.0 - (p.skor_aspd if p and p.skor_aspd else 0.0),
        
        atap=k.jenis_atap_terluas or 0,
        dinding=k.jenis_dinding_terluas or 0,
        lantai=k.jenis_lantai_terluas or 0,
        
        url_foto=f.url_foto if f else None,
        foto_urls=[foto.url_foto for foto in fotos if foto.url_foto],
        visual_match=not p.ada_ketidaksesuaian_visual if p and p.ada_ketidaksesuaian_visual is not None else None,
        visual_reasoning=p.reasoning_tim2 if p else None,
        catatan=p.catatan_petugas if p else None,
        catatan_supervisor=p.catatan_supervisor if p else None,
        
        aiReasoning=p.reasoning_tim3 if p and p.reasoning_tim3 else "Data reasoning belum tersedia dari AI."
    )

@app.put(
    "/api/v1/manajemen-bantuan/{id_keluarga}/status",
    tags=["Manajemen Bantuan"],
    summary="Update status validasi dan rekomendasi bantuan",
    response_model=schemas.DetailKeluargaResponse
)
async def update_status_validasi(
    id_keluarga: UUID,
    request: schemas.UpdateStatusValidasiRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    p = db.query(models.Perhitungan).filter(models.Perhitungan.keluarga_id == id_keluarga).first()
    if not p:
        p = models.Perhitungan(keluarga_id=id_keluarga)
        db.add(p)
    
    if request.status_validasi:
        p.status_validasi = request.status_validasi
        
    if request.bantuan is not None:
        p.rekomendasi_bantuan = request.bantuan
        
    if request.catatan is not None:
        p.catatan_petugas = request.catatan
        
    if request.catatan_supervisor is not None:
        p.catatan_supervisor = request.catatan_supervisor
        
    db.commit()
    return await get_detail_manajemen_bantuan(id_keluarga, current_user, db)

# BAGIAN 11: ENDPOINT READ DATA (GET)
@app.get(
    "/api/v1/keluarga",
    tags=["4. Read Data"],
    summary="Ambil daftar semua keluarga (dengan pagination)"
)
async def list_keluarga(
    skip: int = 0,
    limit: int = 20,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    total = db.query(models.Keluarga).count()
    data = db.query(models.Keluarga).offset(skip).limit(limit).all()
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "data": [schemas.KeluargaResponse.from_orm(k) for k in data]
    }

@app.get(
    "/api/v1/keluarga/{keluarga_id}",
    tags=["4. Read Data"],
    summary="Ambil detail satu keluarga berdasarkan ID"
)
async def get_keluarga(
    keluarga_id: UUID,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    keluarga = db.query(models.Keluarga).filter(models.Keluarga.id == keluarga_id).first()
    if not keluarga:
        raise HTTPException(status_code=404, detail="Keluarga tidak ditemukan")

    foto_terbaru = db.query(models.Foto).filter(
        models.Foto.keluarga_id == keluarga_id
    ).order_by(models.Foto.diunggah_pada.desc()).first()
    
    url_public = foto_terbaru.url_foto if foto_terbaru else None

    return schemas.KeluargaResponse.from_orm(keluarga)

@app.get(
    "/api/v1/keluarga/{keluarga_id}/histori",
    tags=["4. Read Data"],
    summary="Lihat riwayat perubahan asesmen satu keluarga"
)
async def get_histori(
    keluarga_id: UUID,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    logs = db.query(models.LogHistori).filter(
        models.LogHistori.keluarga_id == keluarga_id
    ).order_by(models.LogHistori.timestamp.desc()).all()

    return {
        "keluarga_id": str(keluarga_id),
        "jumlah_riwayat": len(logs),
        "riwayat": [
            {
                "timestamp": log.timestamp,
                "desil_lama": log.desil_lama,
                "desil_baru": log.desil_baru,
                "bantuan_lama": log.bantuan_lama,
                "bantuan_baru": log.bantuan_baru,
            }
            for log in logs
        ]
    }

if __name__ == "__main__":
    print("Menjalankan Main Server di Port 8000...")
    uvicorn.run(app, host="0.0.0.0", port=8000)